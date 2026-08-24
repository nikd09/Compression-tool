"""
Workbook, CSV and HTML output.

The export is where a reader who was not in the room meets the data, so the
tests care as much about what the file explains as about what it contains: the
two stiffness columns must be distinguishable, the permanent-deformation column
must not claim to be compression set, and a thin fit must be visibly flagged.
"""

from __future__ import annotations

import pandas as pd
import pytest
from openpyxl import load_workbook

from compression_tool import Config, diagnostics, ingest
from compression_tool.excel_export import (
    cross_specimen_stats,
    cycles_dataframe,
    summary_pairs,
    write_csv,
    write_workbook,
)
from compression_tool.html_report import render
from compression_tool.persistence import read_json
from compression_tool.schema import (
    QUALITY_MIN_POINTS,
    QUALITY_MIN_R2,
    stiffness_quality,
    user_facing_cycle_columns,
)

from conftest import multistage_signal, write_single_workbook


@pytest.fixture
def payloads(workspace, series_file):
    result = ingest([series_file], workspace, material="PEEK")
    return [read_json(s.json_path) for s in result.specimens]


@pytest.fixture
def single_payload(workspace, single_file):
    result = ingest([single_file], workspace, material="TALCO50")
    return read_json(result.specimens[0].json_path)


# ----------------------------------------------------------------------------
# Workbook structure
# ----------------------------------------------------------------------------


def test_workbook_has_the_expected_sheets(tmp_path, single_payload):
    path = write_workbook([single_payload], tmp_path / "out.xlsx")
    book = load_workbook(path)
    assert book.sheetnames == ["Summary", "Cycles", "Data dictionary", "Config"]


def test_cycle_headers_carry_units(tmp_path, single_payload):
    path = write_workbook([single_payload], tmp_path / "out.xlsx")
    headers = [c.value for c in load_workbook(path)["Cycles"][1]]

    assert "Peak stress (MPa)" in headers
    assert "Stiffness (common band) (MPa/mm)" in headers
    assert "Stiffness (relative band) (MPa/mm)" in headers
    assert "Hold displacement (mm)" in headers
    assert "Maximum displacement (mm)" in headers
    assert "Displacement at peak stress (mm)" in headers
    # Internal bookkeeping stays out of the user-facing table.
    assert not any(str(h).startswith("_") for h in headers if h)


def test_hold_displacement_is_never_separated_from_hold_length(tmp_path, single_payload):
    """Hold displacement without hold length invites reading a longer dwell as
    more movement, which is the exact misreading the pair exists to prevent."""
    from compression_tool.schema import INSEPARABLE_PAIRS

    path = write_workbook([single_payload], tmp_path / "out.xlsx")
    headers = [c.value for c in load_workbook(path)["Cycles"][1] if c.value]
    keys = [c.key for c in user_facing_cycle_columns(has_strain=False)]

    for left, right in INSEPARABLE_PAIRS:
        assert keys.index(right) == keys.index(left) + 1, f"{left} / {right} split"

    assert "Hold length (samples)" in headers
    assert headers.index("Hold displacement (mm)") == headers.index("Hold length (samples)") + 1


def test_per_sample_column_is_not_called_a_rate(tmp_path, single_payload):
    """Samples convert to time only under constant sampling, which the export
    does not record. The column must never present itself as a creep rate."""
    from compression_tool.schema import HOLD_DISP_RATE

    path = write_workbook([single_payload], tmp_path / "out.xlsx")
    headers = [str(c.value) for c in load_workbook(path)["Cycles"][1] if c.value]

    assert "Hold displacement per 1000 samples (mm)" in headers
    assert not any("rate" in h.lower() for h in headers)
    assert "NOT a creep rate" in HOLD_DISP_RATE.description


def test_cycle_rows_match_the_record(tmp_path, single_payload):
    path = write_workbook([single_payload], tmp_path / "out.xlsx")
    sheet = load_workbook(path)["Cycles"]
    headers = [c.value for c in sheet[1]]
    rows = list(sheet.iter_rows(min_row=2, values_only=True))

    assert len(rows) == len(single_payload["cycles"])
    peak_col = headers.index("Peak stress (MPa)")
    for row, cycle in zip(rows, single_payload["cycles"]):
        assert row[peak_col] == pytest.approx(cycle["PeakStress_MPa"])


def test_multi_specimen_workbook_labels_each_row(tmp_path, payloads):
    path = write_workbook(payloads, tmp_path / "run.xlsx")
    sheet = load_workbook(path)["Cycles"]

    assert sheet.cell(row=1, column=1).value == "Specimen"
    labels = {sheet.cell(row=r, column=1).value for r in range(2, sheet.max_row + 1)}
    assert labels == {p["specimen"]["label"] for p in payloads}
    assert sheet.max_row == 1 + sum(len(p["cycles"]) for p in payloads)


def test_summary_compares_specimens_side_by_side(tmp_path, payloads):
    path = write_workbook(payloads, tmp_path / "run.xlsx")
    sheet = load_workbook(path)["Summary"]

    header = [c.value for c in sheet[3]]
    assert header[0] == "Field"
    assert set(header[1:]) == {p["specimen"]["label"] for p in payloads}

    fields = {sheet.cell(row=r, column=1).value for r in range(4, sheet.max_row + 1)}
    assert "Specimen height h0 (mm)" in fields
    assert "Cycles" in fields
    assert "Total permanent deformation (mm)" in fields


def test_data_dictionary_explains_the_traps(tmp_path, single_payload):
    """A reader with only the workbook must be able to tell the two stiffness
    columns apart and must not mistake permanent deformation for compression
    set."""
    path = write_workbook([single_payload], tmp_path / "out.xlsx")
    sheet = load_workbook(path)["Data dictionary"]
    text = "\n".join(
        str(v) for row in sheet.iter_rows(values_only=True) for v in row if v
    )

    assert "IDENTICAL stress window" in text
    assert "NOT comparable across cycles" in text
    assert "ASTM D395" in text
    assert "not seconds" in text


def test_config_sheet_records_the_settings(tmp_path, single_payload):
    path = write_workbook([single_payload], tmp_path / "out.xlsx")
    sheet = load_workbook(path)["Config"]
    pairs = {
        sheet.cell(row=r, column=1).value: sheet.cell(row=r, column=2).value
        for r in range(4, sheet.max_row + 1)
    }

    assert pairs["unload_frac"] == pytest.approx(0.02)
    assert pairs["ref_stress_mpa"] == "auto"
    assert any("reference stress (MPa)" in str(k) for k in pairs)


def test_strain_columns_are_absent_without_h0(tmp_path, single_payload):
    assert single_payload["analysis"]["has_strain"] is False
    path = write_workbook([single_payload], tmp_path / "out.xlsx")
    headers = [c.value for c in load_workbook(path)["Cycles"][1]]

    assert not any("strain" in str(h).lower() for h in headers if h)
    assert not any(str(h).endswith("(%)") for h in headers if h)


def test_strain_columns_are_present_with_h0(tmp_path, payloads):
    assert payloads[0]["analysis"]["has_strain"] is True
    path = write_workbook(payloads, tmp_path / "run.xlsx")
    headers = [c.value for c in load_workbook(path)["Cycles"][1]]

    assert "Strain at peak stress (%)" in headers


# ----------------------------------------------------------------------------
# Quality flag
# ----------------------------------------------------------------------------


@pytest.mark.parametrize(
    "n,r2,expected",
    [
        (200, 0.999, "ok"),
        (QUALITY_MIN_POINTS - 1, 0.999, "few points"),
        (200, QUALITY_MIN_R2 - 0.01, "nonlinear"),
        (None, None, "none"),
        (3, 0.10, "few points"),
    ],
)
def test_stiffness_quality_flag(n, r2, expected):
    assert stiffness_quality(n, r2) == expected


def test_quality_column_sits_next_to_the_fit_it_describes():
    cols = [c.key for c in user_facing_cycle_columns(has_strain=False)]
    assert cols.index("Stiffness_common_quality") == cols.index("Stiffness_common_r2") + 1


def test_workbook_carries_the_quality_column(tmp_path, single_payload):
    path = write_workbook([single_payload], tmp_path / "out.xlsx")
    sheet = load_workbook(path)["Cycles"]
    headers = [c.value for c in sheet[1]]
    col = headers.index("Stiffness (common band), quality") + 1

    values = {sheet.cell(row=r, column=col).value for r in range(2, sheet.max_row + 1)}
    assert values <= {"ok", "few points", "nonlinear", "none"}


# ----------------------------------------------------------------------------
# CSV
# ----------------------------------------------------------------------------


def test_csv_matches_the_workbook_columns(tmp_path, single_payload):
    csv_path = write_csv([single_payload], tmp_path / "out.csv")
    frame = pd.read_csv(csv_path)
    headers = [c.value for c in load_workbook(
        write_workbook([single_payload], tmp_path / "out.xlsx"))["Cycles"][1]]

    assert list(frame.columns) == [h for h in headers if h]
    assert len(frame) == len(single_payload["cycles"])


def test_csv_can_carry_a_specimen_column(tmp_path, payloads):
    frame = cycles_dataframe(payloads, with_specimen=True)
    assert frame.columns[0] == "Specimen"
    assert len(frame) == sum(len(p["cycles"]) for p in payloads)


# ----------------------------------------------------------------------------
# HTML
# ----------------------------------------------------------------------------


def test_html_is_self_contained(single_payload):
    page = render([single_payload])
    assert page.startswith("<!doctype html>")
    assert "<style>" in page
    for external in ("http://", "https://", "<script"):
        assert external not in page


def test_html_warns_about_multi_stage_and_missing_h0(payloads, single_payload):
    multi = render(payloads)
    assert "Multi-stage test" in multi
    assert "not comparable across stages" in multi

    bare = render([single_payload])
    assert "strain-normalised columns are omitted" in bare


def test_html_surfaces_loader_notes(single_payload):
    page = render([single_payload])
    assert "displacement channels found" in page


def test_html_escapes_content(single_payload):
    hostile = dict(single_payload)
    hostile["specimen"] = {**single_payload["specimen"], "label": "<img src=x onerror=1>"}
    page = render([hostile])

    assert "<img src=x" not in page


# ----------------------------------------------------------------------------
# Source filename vs ingest-machine path
# ----------------------------------------------------------------------------


def test_source_filename_is_the_basename_not_the_ingest_path(single_payload, single_file):
    """'Source file' in the Summary must read as the original filename, not
    wherever the ingest happened to run from -- a sandbox temp path is not
    meaningful to an operator on a different machine."""
    spec = single_payload["specimen"]
    assert spec["source_filename"] == single_file.name
    assert "/" not in spec["source_filename"]
    # The full path is still kept, under a name that says what it is.
    assert spec["source_file"].endswith(single_file.name)


def test_workbook_labels_filename_and_path_separately(tmp_path, single_payload):
    path = write_workbook([single_payload], tmp_path / "out.xlsx")
    sheet = load_workbook(path)["Summary"]
    labels = {sheet.cell(row=r, column=1).value for r in range(1, sheet.max_row + 1)}

    assert "Source file" in labels
    assert "Source path (ingest machine)" in labels


# ----------------------------------------------------------------------------
# Hysteresis loss: multi-stage scoping
# ----------------------------------------------------------------------------


def test_mean_hysteresis_label_is_scoped_for_multi_stage(payloads):
    """Multi-stage cycles span different stress levels, and hysteresis loss is
    not flat across a stress range, so an unscoped 'Mean hysteresis loss'
    would read as a single physical value when it is not one."""
    from compression_tool.excel_export import summary_pairs

    assert payloads[0]["analysis"]["multi_stage"] is True
    keys = [k for k, _ in summary_pairs(payloads[0])]
    assert any("Mean hysteresis loss across cycles" in k for k in keys)
    assert not any(k == "Mean hysteresis loss (-)" for k in keys)


def test_mean_hysteresis_label_is_plain_for_constant_amplitude(tmp_path):
    """A constant-amplitude test has one stress level, so the mean IS a
    single physical value and does not need the multi-stage caveat."""
    stress, disp = multistage_signal(stages=(300.0, 300.0, 300.0))
    path = write_single_workbook(tmp_path / "flat.xlsx", stress, disp, disp)
    result = ingest([path], tmp_path / "ws", material="flat", cfg=Config())
    payload = read_json(result.specimens[0].json_path)

    assert payload["analysis"]["multi_stage"] is False
    keys = [k for k, _ in summary_pairs(payload)]
    assert "Mean hysteresis loss (-)" in keys
    assert not any("across cycles" in k for k in keys)


# ----------------------------------------------------------------------------
# Warnings shown once, not once per specimen column
# ----------------------------------------------------------------------------


def test_summary_sheet_shows_each_warning_once(tmp_path, payloads):
    """Both specimens of the series were ingested under the same config and
    trip the same warnings; the Summary sheet must show each paragraph once,
    not once per specimen column."""
    expected = diagnostics.distinct(payloads)
    assert expected  # the fixture should trip at least the gauge-length warning

    path = write_workbook(payloads, tmp_path / "run.xlsx")
    sheet = load_workbook(path)["Summary"]
    severities = [sheet.cell(row=r, column=1).value for r in range(1, sheet.max_row + 1)
                 if sheet.cell(row=r, column=1).value in ("CRITICAL", "CAUTION", "INFO")]
    assert len(severities) == len(expected)


def test_html_shows_each_warning_once(payloads):
    import html as html_module

    expected = diagnostics.distinct(payloads)
    page = render(payloads)
    for w in expected:
        # render() HTML-escapes the message (quotes -> entities), so compare
        # against the same escaping rather than the raw text.
        assert page.count(html_module.escape(w["message"])) == 1


# ----------------------------------------------------------------------------
# Cross-specimen statistics
# ----------------------------------------------------------------------------


def test_cross_specimen_stats_needs_more_than_one_specimen(single_payload):
    assert cross_specimen_stats([single_payload]) == []


def test_cross_specimen_stats_covers_every_cycle(payloads):
    stats = cross_specimen_stats(payloads)
    assert stats
    peak_stress = next(e for e in stats if e["key"] == "PeakStress_MPa")
    assert [r["cycle"] for r in peak_stress["rows"]] == list(range(1, 10))
    # Both specimens carry the same commanded peak, so agreement is tight.
    assert all(r["cov_pct"] is not None and r["cov_pct"] < 5 for r in peak_stress["rows"])


def test_statistics_sheet_only_appears_with_multiple_specimens(tmp_path, single_payload, payloads):
    single_path = write_workbook([single_payload], tmp_path / "single.xlsx")
    assert "Statistics" not in load_workbook(single_path).sheetnames

    multi_path = write_workbook(payloads, tmp_path / "multi.xlsx")
    assert "Statistics" in load_workbook(multi_path).sheetnames


def test_html_statistics_section_only_appears_with_multiple_specimens(single_payload, payloads):
    assert "Cross-specimen statistics" not in render([single_payload])
    assert "Cross-specimen statistics" in render(payloads)


# ----------------------------------------------------------------------------
# Four specimens in one export
# ----------------------------------------------------------------------------


@pytest.fixture
def quad_payloads(workspace, four_specimen_file):
    result = ingest([four_specimen_file], workspace, material="QUAD")
    return [read_json(s.json_path) for s in result.specimens]


def test_four_specimens_all_reach_the_workbook(tmp_path, quad_payloads):
    assert len(quad_payloads) == 4
    path = write_workbook(quad_payloads, tmp_path / "quad.xlsx")
    book = load_workbook(path)
    assert "Statistics" in book.sheetnames
    # Summary lays specimens across the page: one label column plus one per
    # specimen, so a dropped specimen shows up as a missing column here.
    summary = book["Summary"]
    assert summary.max_column >= 5


def test_cross_specimen_stats_averages_over_all_four(quad_payloads):
    stats = cross_specimen_stats(quad_payloads)
    peak = next(e for e in stats if e["key"] == "PeakStress_MPa")
    assert [r["cycle"] for r in peak["rows"]] == list(range(1, 10))
    assert all(r["n"] == 4 for r in peak["rows"])


def test_four_specimen_csv_keeps_every_row(tmp_path, quad_payloads):
    path = write_csv(quad_payloads, tmp_path / "quad.csv", with_specimen=True)
    df = pd.read_csv(path)
    assert len(df) == sum(len(p["cycles"]) for p in quad_payloads)
    assert df["Specimen"].nunique() == 4


# ----------------------------------------------------------------------------
# Combined-run report title
# ----------------------------------------------------------------------------


def test_run_report_title_does_not_repeat_the_material(workspace, series_file):
    """title=f'{material} - {run_dir.name}' repeated the material, since
    run_dir.name already starts with the material slug: 'T050E1 -
    T050E1_2026-08-17'."""
    result = ingest([series_file], workspace, material="T050E1")
    title_line = result.run_html.read_text(encoding="utf-8")
    start = title_line.index("<title>") + len("<title>")
    end = title_line.index("</title>")
    title = title_line[start:end]

    assert title.count("T050E1") == 1
    assert "T050E1 -" in title


# ----------------------------------------------------------------------------
# Files written by the pipeline
# ----------------------------------------------------------------------------


def test_every_artifact_is_written(workspace, series_file):
    result = ingest([series_file], workspace, material="PEEK")

    for specimen in result.specimens:
        for path in (specimen.json_path, specimen.csv_path,
                     specimen.xlsx_path, specimen.html_path):
            assert path.exists() and path.stat().st_size > 0

    # Two specimens in the export, so a combined run workbook is written too.
    assert result.run_xlsx is not None and result.run_xlsx.exists()
    assert result.run_html is not None and result.run_html.exists()
    assert result.run_xlsx.with_suffix(".csv").exists()


def test_single_specimen_run_does_not_duplicate_the_workbook(workspace, single_file):
    result = ingest([single_file], workspace, material="TALCO50")
    assert result.run_xlsx == result.specimens[0].xlsx_path
    assert len(list(result.run_dir.glob("*.xlsx"))) == 1
