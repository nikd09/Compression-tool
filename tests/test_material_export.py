"""
material_export.py: the combined-across-runs workbook and dashboard.

A material is normally built up across several separate ingest sessions, not
one. These tests exist to pin the one thing that matters about that: the
combined export always covers every specimen ever ingested for a material,
not just whichever run happened to trigger the rebuild.
"""

from __future__ import annotations

from openpyxl import load_workbook

from compression_tool import Workspace, export_material, ingest


def test_combines_specimens_from_separate_ingest_sessions(workspace, series_file, single_file):
    """Two specimens land in one run; a third lands in a second, later run.
    The rollup must show all three, not just the run that last triggered it."""
    ingest([series_file], workspace, material="PEEK")
    ingest([single_file], workspace, material="PEEK")

    ws = Workspace.at(workspace)
    result = export_material(ws, "PEEK")
    assert result["xlsx"] is not None and result["xlsx"].exists()
    assert result["html"] is not None and result["html"].exists()

    book = load_workbook(result["xlsx"], read_only=True)
    header = next(book["Summary"].iter_rows(min_row=3, max_row=3, values_only=True))
    # First cell is the "Field" label column; the rest are one per specimen.
    specimen_columns = [c for c in header[1:] if c is not None]
    assert len(specimen_columns) == 3

    html = result["html"].read_text(encoding="utf-8")
    assert "/*__DATA__*/" not in html  # the template placeholder was replaced
    assert html.count('"label"') == 3  # one dashboard specimen block each


def test_ingest_triggers_the_rebuild_automatically(workspace, series_file, single_file):
    """The whole point is that nobody has to remember to run this by hand."""
    ingest([series_file], workspace, material="PEEK")
    result = ingest([single_file], workspace, material="PEEK")

    assert result.material_xlsx is not None
    assert result.material_xlsx.exists()
    book = load_workbook(result.material_xlsx, read_only=True)
    header = next(book["Summary"].iter_rows(min_row=3, max_row=3, values_only=True))
    assert len([c for c in header[1:] if c is not None]) == 3


def test_a_different_material_is_not_pulled_in(workspace, series_file, single_file):
    ingest([series_file], workspace, material="PEEK")
    ingest([single_file], workspace, material="TALCO50")

    ws = Workspace.at(workspace)
    result = export_material(ws, "PEEK")
    book = load_workbook(result["xlsx"], read_only=True)
    header = next(book["Summary"].iter_rows(min_row=3, max_row=3, values_only=True))
    assert len([c for c in header[1:] if c is not None]) == 2


def test_no_specimens_returns_none_rather_than_an_empty_file(workspace):
    ws = Workspace.at(workspace)
    result = export_material(ws, "NothingIngestedYet")
    assert result == {"xlsx": None, "html": None}


def test_html_title_names_the_material_not_one_specimens_file(workspace, series_file):
    ingest([series_file], workspace, material="PEEK-GF30")
    ws = Workspace.at(workspace)
    result = export_material(ws, "PEEK-GF30")
    html = result["html"].read_text(encoding="utf-8")
    assert "<title>PEEK-GF30 - Compression Results</title>" in html


def test_survives_a_specimen_record_deleted_outside_the_app(
    workspace, series_file, single_file
):
    """The index can still list a specimen whose JSON record was deleted
    straight from disk (Explorer, the shared drive) rather than through
    the app -- nothing but a reindex clears it out. Building the combined
    export must skip that one record, not crash and take the whole
    material's report down with it."""
    ingest([series_file], workspace, material="PEEK")
    result = ingest([single_file], workspace, material="PEEK")

    # Delete one specimen's record straight off disk, index untouched --
    # exactly what happens when someone removes a file from Explorer.
    stale_json = next(iter(result.specimens)).json_path
    stale_json.unlink()

    ws = Workspace.at(workspace)
    exported = export_material(ws, "PEEK")
    assert exported["xlsx"] is not None and exported["xlsx"].exists()

    book = load_workbook(exported["xlsx"], read_only=True)
    header = next(book["Summary"].iter_rows(min_row=3, max_row=3, values_only=True))
    # Only the 2 specimens whose records still exist -- 3 ingested, 1 deleted.
    assert len([c for c in header[1:] if c is not None]) == 2
