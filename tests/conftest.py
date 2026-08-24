"""
Synthetic Zwick exports.

The real export files are not in the repository, so the suite builds signals
that reproduce the behaviours the sample data actually showed -- rising stage
peaks, a dwell at peak during which displacement keeps climbing, and a collapse
to a few-micrometre baseline at near-zero stress. Those three are what the
engine's less obvious choices exist for, so a synthetic signal without them
would test nothing that matters.

Real-file regression pins live in test_regression_real_files.py and activate as
soon as the exports are dropped into tests/data/.
"""

from __future__ import annotations

import sys
from pathlib import Path

import numpy as np
import pytest
from openpyxl import Workbook

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

# Contact is lost below a few MPa and the signal falls back to this baseline.
BASELINE_MM = 0.003
# Stress scale over which contact re-establishes on the loading branch.
CONTACT_SCALE_MPA = 2.0

STAGES = (50.0, 100.0, 150.0, 200.0, 250.0, 300.0, 350.0, 400.0, 450.0)
H0_MM = 0.471
D0_MM = 16.0
TEMPERATURE_C = 23.0


# ----------------------------------------------------------------------------
# Signal generation
# ----------------------------------------------------------------------------


def _contact(stress: np.ndarray) -> np.ndarray:
    return 1.0 - np.exp(-stress / CONTACT_SCALE_MPA)


def cycle_arrays(
    peak: float,
    x_perm: float,
    amplitude: float,
    creep: float,
    set_inc: float,
    *,
    n_ramp: int = 400,
    n_hold: int = 1000,
    n_unload: int = 400,
    n_rest: int = 300,
    hold: bool = True,
) -> tuple[np.ndarray, np.ndarray]:
    """One load-dwell-unload-rest cycle.

    The unloading branch is offset to the right of the loading branch by the
    permanent set gained in this cycle, which is what makes the enclosed loop
    area -- the dissipated energy -- positive.
    """
    x_res = x_perm + set_inc

    s_load = np.linspace(0.0, peak, n_ramp)
    x_load = (
        BASELINE_MM
        + (x_perm - BASELINE_MM) * _contact(s_load)
        + amplitude * (s_load / peak) ** 0.85
    )

    if hold and n_hold > 0:
        t = np.linspace(0.0, 1.0, n_hold)
        # Ripple stays inside Config.hold_tol_frac so the plateau is detected
        # as one dwell rather than several.
        s_hold = peak * (1.0 + 0.0005 * np.sin(2 * np.pi * 3 * t))
        x_hold = x_load[-1] + creep * np.sqrt(t)
    else:
        s_hold = np.empty(0)
        x_hold = np.empty(0)

    x_max = float(x_hold[-1]) if x_hold.size else float(x_load[-1])

    s_unload = np.linspace(peak, 0.0, n_unload)
    x_unload = (
        BASELINE_MM
        + (x_res - BASELINE_MM) * _contact(s_unload)
        + (x_max - x_res) * (s_unload / peak) ** 0.85
    )

    # Second dwell near zero: contact is lost, displacement drops to baseline.
    s_rest = np.full(n_rest, 0.0005 * peak)
    x_rest = np.full(n_rest, BASELINE_MM)

    return (
        np.concatenate([s_load, s_hold, s_unload, s_rest]),
        np.concatenate([x_load, x_hold, x_unload, x_rest]),
    )


def multistage_signal(
    stages: tuple[float, ...] = STAGES, *, hold: bool = True
) -> tuple[np.ndarray, np.ndarray]:
    """A rising multi-stage test: peak stress climbs every cycle."""
    top = max(stages)
    x_perm = BASELINE_MM
    stress_parts = [np.zeros(200)]
    disp_parts = [np.full(200, BASELINE_MM)]

    for peak in stages:
        amplitude = 0.05 * (peak / top) ** 0.85
        creep = 0.004 * (peak / top)
        set_inc = 0.0015 * (peak / top)
        s, x = cycle_arrays(peak, x_perm, amplitude, creep, set_inc, hold=hold)
        stress_parts.append(s)
        disp_parts.append(x)
        x_perm += set_inc

    return np.concatenate(stress_parts), np.concatenate(disp_parts)


def expected_permanent_set(
    stages: tuple[float, ...] = STAGES, residual_stress_mpa: float | None = None
) -> float:
    """Permanent set accumulated between the first and last cycle, in mm.

    Available in closed form because the elastic contribution cancels: the
    amplitude scales with the stage peak exactly as the reference-stress
    fraction shrinks, so the part read at the residual reference stress is
    identical in every stage and only the accumulated set survives.
    """
    top = max(stages)
    total = sum(0.0015 * (p / top) for p in stages[:-1])
    if residual_stress_mpa is None:
        return total
    return total * float(1.0 - np.exp(-residual_stress_mpa / CONTACT_SCALE_MPA))


# ----------------------------------------------------------------------------
# Workbook writers
# ----------------------------------------------------------------------------


def write_single_workbook(
    path: Path,
    stress: np.ndarray,
    disp_extenso: np.ndarray,
    disp_crosshead: np.ndarray,
) -> Path:
    """One sample per sheet, two displacement channels, displacement in mm."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Serie 1"

    ws.append(["Zwick Z100 - Druckversuch", None, None])
    ws.append(["Standardweg", "Sonder LAA", "Druckspannung"])
    ws.append(["mm", "mm", "N/mm²"])
    for c, e, s in zip(disp_crosshead, disp_extenso, stress):
        ws.append([float(c), float(e), float(s)])

    wb.save(path)
    return path


def write_series_workbook(
    path: Path,
    specimens: dict[str, tuple[np.ndarray, np.ndarray]],
    *,
    with_metadata: bool = True,
    disp_unit: str = "µm",
) -> Path:
    """Multi-sample workbook: sample columns side by side, displacement in µm.

    The values sheet is named with a trailing space, as the real exports are.
    """
    wb = Workbook()
    values = wb.active
    values.title = "Werte Serie "

    ids_row: list = []
    chan_row: list = []
    unit_row: list = []
    for sid in specimens:
        ids_row += [sid, sid]
        chan_row += ["Standardweg", "Druckspannung"]
        unit_row += [disp_unit, "MPa"]
    values.append(ids_row)
    values.append(chan_row)
    values.append(unit_row)

    scale = 1000.0 if disp_unit in ("µm", "um") else 1.0
    length = max(len(s) for s, _ in specimens.values())
    for i in range(length):
        row: list = []
        for stress, disp in specimens.values():
            row.append(float(disp[i]) * scale if i < len(disp) else None)
            row.append(float(stress[i]) if i < len(stress) else None)
        values.append(row)

    if with_metadata:
        meta = wb.create_sheet("Ergebnisse Serie ")
        meta.append(["Probe", "h0 in mm", "d0 in mm", "Temperatur in °C"])
        meta.append(["Nr.", "mm", "mm", "°C"])
        for sid in specimens:
            meta.append([sid, H0_MM, D0_MM, TEMPERATURE_C])

    wb.save(path)
    return path


# ----------------------------------------------------------------------------
# Fixtures
# ----------------------------------------------------------------------------


@pytest.fixture(scope="session")
def signal() -> tuple[np.ndarray, np.ndarray]:
    return multistage_signal()


@pytest.fixture(scope="session")
def single_file(tmp_path_factory, signal) -> Path:
    stress, disp = signal
    # The crosshead channel reads higher: it includes machine compliance.
    crosshead = disp * 1.35 + 0.01
    path = tmp_path_factory.mktemp("exports") / "TALCO50.xlsx"
    return write_single_workbook(path, stress, disp, crosshead)


@pytest.fixture(scope="session")
def series_file(tmp_path_factory, signal) -> Path:
    stress, disp = signal
    # A second specimen of the same series, marginally softer.
    specimens = {"1": (stress, disp), "2": (stress, disp * 1.04)}
    path = tmp_path_factory.mktemp("exports") / "Mehrstufiger.xlsx"
    return write_series_workbook(path, specimens)


@pytest.fixture(scope="session")
def four_specimen_file(tmp_path_factory, signal) -> Path:
    """A four-sample series in one export.

    Real tests are not always run in pairs, and every stage of the pipeline --
    segmentation, the cross-specimen statistics sheet, the dashboard's colour
    slots -- has to hold when the count is neither one nor two. Each specimen
    is scaled slightly differently so they stay individually identifiable.
    """
    stress, disp = signal
    specimens = {
        "1": (stress, disp),
        "2": (stress, disp * 1.04),
        "3": (stress, disp * 0.97),
        "4": (stress, disp * 1.09),
    }
    path = tmp_path_factory.mktemp("exports") / "VierProben.xlsx"
    return write_series_workbook(path, specimens)


@pytest.fixture
def workspace(tmp_path) -> Path:
    return tmp_path / "data"
