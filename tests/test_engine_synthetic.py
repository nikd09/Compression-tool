"""
Engine behaviour on synthetic exports.

These lock down the findings that shaped the engine. Each test names the
failure it prevents, because the value of pinning e.g. non-negative dissipated
energy is entirely in remembering why it could ever have been negative.
"""

from __future__ import annotations

import numpy as np
import pytest

from compression_tool.core import (
    Config,
    TestData,
    analyse_test,
    detect_format,
    load_tests,
    segment_cycles,
)

from conftest import (
    BASELINE_MM,
    D0_MM,
    H0_MM,
    STAGES,
    TEMPERATURE_C,
    expected_permanent_set_incremental,
    multistage_signal,
    write_series_workbook,
    write_single_workbook,
)


# ----------------------------------------------------------------------------
# Format detection and loading
# ----------------------------------------------------------------------------


def test_detect_format_distinguishes_the_two_layouts(single_file, series_file):
    assert detect_format(str(single_file)) == "single"
    assert detect_format(str(series_file)) == "series"


def test_single_format_prefers_the_extensometer_channel(single_file):
    """The crosshead signal includes machine compliance, so picking it would
    understate stiffness. The choice must come from the channel name."""
    (test,) = load_tests(str(single_file))

    assert test.displacement_channel == "Sonder LAA"
    assert any("2 displacement channels" in n for n in test.notes)
    # The crosshead channel was scaled by 1.35 in the fixture; the loaded
    # signal must not carry that inflation.
    assert float(np.max(test.displacement_mm)) < 0.07


def test_series_format_reads_every_specimen_and_its_metadata(series_file):
    tests = load_tests(str(series_file))

    assert len(tests) == 2
    assert {t.label.split("_S")[-1] for t in tests} == {"1", "2"}
    for test in tests:
        assert test.h0_mm == pytest.approx(H0_MM)
        assert test.d0_mm == pytest.approx(D0_MM)
        assert test.temperature_c == pytest.approx(TEMPERATURE_C)


def test_h0_column_is_not_fooled_by_a_percentage_column_named_the_same(tmp_path, signal):
    """A results sheet can carry a relative-deviation column like "dh0/h0 in
    %" alongside the real "h0 in mm" measurement. Both contain the substring
    "h0"; picking whichever comes first would silently bind h0_mm to a
    percentage -- wrong strain and modulus, with no warning. The real, mm
    column must win regardless of column order."""
    stress, disp = signal
    path = write_series_workbook(
        tmp_path / "ambiguous.xlsx", {"1": (stress, disp)}, with_metadata=False
    )

    from openpyxl import load_workbook

    wb = load_workbook(path)
    meta = wb.create_sheet("Ergebnisse Serie ")
    # The deviation column is placed FIRST, exactly the ordering that broke
    # a first-substring-match resolver.
    meta.append(["Probe", "dh0/h0 in %", "h0 in mm", "d0 in mm", "Temperatur in °C"])
    meta.append(["Nr.", "%", "mm", "mm", "°C"])
    meta.append(["1", 2.5, H0_MM, D0_MM, TEMPERATURE_C])
    wb.save(path)

    (test,) = load_tests(str(path))
    assert test.h0_mm == pytest.approx(H0_MM)
    assert test.d0_mm == pytest.approx(D0_MM)


def test_micrometre_and_millimetre_exports_agree(tmp_path, signal):
    """Units are parsed from the header row, so the same physical test written
    in µm and in mm must analyse identically."""
    stress, disp = signal
    um = write_series_workbook(tmp_path / "um.xlsx", {"1": (stress, disp)}, disp_unit="µm")
    mm = write_series_workbook(tmp_path / "mm.xlsx", {"1": (stress, disp)}, disp_unit="mm")

    (a,), (b,) = load_tests(str(um)), load_tests(str(mm))
    np.testing.assert_allclose(a.displacement_mm, b.displacement_mm, rtol=1e-9)


def test_stress_unit_synonyms_are_equivalent(tmp_path, signal):
    stress, disp = signal
    mpa = write_series_workbook(tmp_path / "a.xlsx", {"1": (stress, disp)}, disp_unit="mm")
    (a,) = load_tests(str(mpa))

    # The single-format fixture writes N/mm², which must scale identically.
    b_path = write_single_workbook(tmp_path / "b.xlsx", stress, disp, disp)
    (b,) = load_tests(str(b_path))
    np.testing.assert_allclose(a.stress_mpa, b.stress_mpa, rtol=1e-9)


# ----------------------------------------------------------------------------
# Segmentation and holds
# ----------------------------------------------------------------------------


def test_segmentation_finds_every_stage(signal):
    stress, _ = signal
    assert len(segment_cycles(stress, Config())) == len(STAGES)


def test_thresholds_are_relative_not_absolute():
    """The old tool's 5 MPa floor would have discarded every cycle of a low
    stress test. Scaling the whole signal must not change the cycle count."""
    stress, disp = multistage_signal(stages=(0.5, 1.0, 1.5, 2.0))
    assert len(segment_cycles(stress, Config())) == 4

    big, _ = multistage_signal(stages=(500.0, 1000.0, 1500.0, 2000.0))
    assert len(segment_cycles(big, Config())) == 4


def test_cycles_expand_into_the_valleys(signal):
    """A cycle that began above the detection threshold could never be
    interpolated at a reference stress below it, so the residual readout would
    silently go missing."""
    stress, _ = signal
    cfg = Config()
    residual = cfg.residual_stress_frac * float(np.nanmax(stress))

    for start, end in segment_cycles(stress, cfg):
        assert float(stress[start]) < residual


def test_hold_is_detected_in_every_cycle(signal):
    stress, disp = signal
    df = analyse_test(_as_test(stress, disp))
    assert df["HoldDetected"].all()
    assert (df["HoldPoints"] > 900).all()


def test_creep_is_omitted_not_zeroed_when_there_is_no_dwell():
    """A cycle without a dwell must report no creep, not a creep of zero --
    the two mean different things to a reader."""
    stress, disp = multistage_signal(hold=False)
    df = analyse_test(_as_test(stress, disp))

    assert not df["HoldDetected"].any()
    assert df["Creep_during_hold_mm"].isna().all()


# ----------------------------------------------------------------------------
# Metrics
# ----------------------------------------------------------------------------


def test_dissipated_energy_is_never_negative(signal):
    """Splitting the loop at maximum stress rather than maximum displacement
    drives this negative, which is physically impossible."""
    stress, disp = signal
    df = analyse_test(_as_test(stress, disp))

    assert (df["Energy_dissipated_MPa_mm"] > 0).all()
    assert (df["HysteresisLoss_rel"] > 0).all()
    assert (df["HysteresisLoss_rel"] < 1).all()


def test_displacement_peaks_after_stress_does(signal):
    """The specimen keeps creeping while stress is already held constant."""
    stress, disp = signal
    cfg = Config()
    for start, end in segment_cycles(stress, cfg):
        s, x = stress[start : end + 1], disp[start : end + 1]
        assert int(np.argmax(x)) > int(np.argmax(s))


def test_stress_at_max_displacement_equals_peak_on_an_intact_specimen(signal):
    """The synthetic signal stops creeping when the dwell ends, so maximum
    displacement lands at peak stress. A value below the peak means the
    specimen went on compacting while the load was coming off -- the damage
    signature this column exists to catch."""
    stress, disp = signal
    df = analyse_test(_as_test(stress, disp))

    ratio = df["StressAtMaxDisp_MPa"] / df["PeakStress_MPa"]
    assert (ratio > 0.999).all()


def test_stress_at_max_displacement_detects_yielding_through_unload():
    """A specimen that keeps compacting as the load is removed puts its maximum
    displacement partway down the unloading ramp, not at the dwell's end."""
    from conftest import BASELINE_MM, CONTACT_SCALE_MPA

    peak, n_ramp, n_hold, n_unload = 200.0, 400, 600, 400
    s_load = np.linspace(0.0, peak, n_ramp)
    contact = 1.0 - np.exp(-s_load / CONTACT_SCALE_MPA)
    x_load = BASELINE_MM + (BASELINE_MM - BASELINE_MM) * contact + 0.02 * (s_load / peak) ** 0.85
    s_hold = np.full(n_hold, peak)
    x_hold = x_load[-1] + 0.002 * np.linspace(0, 1, n_hold) ** 0.5
    # Unloading: displacement keeps rising over the first third of the ramp
    # before the specimen finally recovers -- continued yielding under falling load.
    s_unload = np.linspace(peak, 0.0, n_unload)
    frac = np.linspace(0, 1, n_unload)
    x_unload = x_hold[-1] + 0.004 * np.sin(np.pi * np.clip(frac / 0.6, 0, 1)) - 0.01 * frac**2
    stress = np.concatenate([np.zeros(200), s_load, s_hold, s_unload, np.full(200, 0.001)])
    disp = np.concatenate([np.full(200, BASELINE_MM), x_load, x_hold, x_unload,
                           np.full(200, BASELINE_MM)])

    df = analyse_test(_as_test(stress, disp))
    assert len(df) == 1
    row = df.iloc[0]
    assert row["StressAtMaxDisp_MPa"] < row["PeakStress_MPa"] * 0.95
    # The maximum is still the maximum, whichever branch it fell on.
    assert row["MaxDisp_mm"] == pytest.approx(float(np.max(disp)))
    # And the energy split still yields non-negative dissipation.
    assert row["Energy_dissipated_MPa_mm"] >= 0


def test_common_band_stiffness_is_comparable_across_stages(signal):
    """The whole point of the common band: an identical stress window in every
    cycle, so a rising stage does not masquerade as a stiffening material."""
    stress, disp = signal
    df = analyse_test(_as_test(stress, disp))

    common = df["Stiffness_common_MPa_per_mm"]
    assert common.notna().all()
    assert common.std() / common.mean() < 0.01

    # The relative band, by contrast, is measured further up the curve on each
    # successive stage, so it rises artificially even though the material has
    # not stiffened. It must NOT be used for cross-stage comparison.
    relative = df["Stiffness_relative_MPa_per_mm"]
    assert relative.is_monotonic_increasing
    assert relative.iloc[-1] > relative.iloc[0] * 1.1


def test_stiffness_reports_its_own_fit_quality(signal):
    """A slope from a handful of points is not trustworthy; n and R2 exist so
    the UI can say so rather than plot it as solid."""
    stress, disp = signal
    df = analyse_test(_as_test(stress, disp))

    assert (df["Stiffness_common_n"] >= 3).all()
    assert (df["Stiffness_common_r2"] > 0.99).all()
    # Later stages cross the fixed window in fewer samples.
    assert df["Stiffness_common_n"].iloc[-1] < df["Stiffness_common_n"].iloc[0]


def test_permanent_deformation_accumulates_to_the_expected_value(signal):
    """Pinned against the closed form of the synthetic signal, so a change in
    how the residual is read shows up as a number rather than a vibe.

    PermDef_incremental_mm is now WITHIN-cycle (ResidualDisp_unload_mm minus
    ResidualDisp_mm, both read in the SAME cycle) rather than referenced to
    cycle 1 -- see core.analyse_test. Cycle 1 therefore has a real,
    well-defined incremental value of its own instead of being undefined by
    construction; PermDef_cumulative_mm is the running total of those."""
    stress, disp = signal
    df = analyse_test(_as_test(stress, disp))

    cumulative = df["PermDef_cumulative_mm"]
    incremental = df["PermDef_incremental_mm"]
    assert cumulative.is_monotonic_increasing
    assert (incremental > 0).all()  # every stage gains SOME permanent set

    residual_stress = df.attrs["residual_stress_mpa"]
    expected_incremental = expected_permanent_set_incremental(STAGES, residual_stress)
    np.testing.assert_allclose(incremental.to_numpy(), expected_incremental, rtol=2e-3)
    assert cumulative.iloc[-1] == pytest.approx(expected_incremental.sum(), rel=2e-3)


def test_residual_is_read_above_the_contact_loss_baseline(signal):
    """Reading permanent set at zero stress would return the unloaded baseline
    of a few micrometres and mean nothing. True on both branches."""
    stress, disp = signal
    df = analyse_test(_as_test(stress, disp))
    assert (df["ResidualDisp_mm"] > BASELINE_MM).all()
    assert (df["ResidualDisp_unload_mm"] > BASELINE_MM).all()


def test_residual_stress_is_reachable_in_the_smallest_cycle(signal):
    """residual_stress is the one reference stress now, anchored to the
    GLOBAL peak (not the smallest cycle's own peak, the way the old,
    now-removed ref_stress was -- see Config). At its low default fraction
    (0.02) it still lands well inside even the smallest stage of a
    realistic multi-stage test, so both branches stay readable at every
    cycle."""
    stress, disp = signal
    df = analyse_test(_as_test(stress, disp))

    assert df.attrs["residual_stress_mpa"] < df["PeakStress_MPa"].min()
    assert df["ResidualDisp_mm"].notna().all()
    assert df["ResidualDisp_unload_mm"].notna().all()
    # The loop is open at the reference stress: unloading sits to the right.
    assert (df["ResidualDisp_unload_mm"] > df["ResidualDisp_mm"]).all()


def test_single_cycle_test_reports_real_permanent_deformation():
    """The OLD formula referenced PermDef_cumulative_mm to cycle 1's own
    loading-branch reading -- for a single-cycle test that is the cycle
    comparing itself to itself, always exactly 0.0 regardless of how much
    permanent set actually occurred. The within-cycle redefinition
    (ResidualDisp_unload_mm minus ResidualDisp_mm, both read in this one
    cycle) reports the real, non-zero value the OLD formula could never
    see -- this is the scientific-validity gap the redesign exists to
    close for a genuinely single-cycle compression test."""
    from conftest import BASELINE_MM, cycle_arrays

    stress, disp = cycle_arrays(
        peak=300.0, x_perm=BASELINE_MM, amplitude=0.05, creep=0.004, set_inc=0.0015
    )
    df = analyse_test(_as_test(stress, disp))

    assert len(df) == 1
    incremental = df["PermDef_incremental_mm"].iloc[0]
    assert incremental == pytest.approx(0.0015, rel=0.2)  # the set_inc gained
    assert df["PermDef_cumulative_mm"].iloc[0] == pytest.approx(incremental)


def test_a_valley_that_only_partially_relaxes_still_separates_cycles():
    """Reproduces (synthetically) the first real failure this redesign was
    built to fix: MeshG_3mpa_10cyc_4.xlsx's cycle 1/cycle 2 boundary only
    relaxed to ~0.85 MPa, not near zero, and the OLD absolute unload_frac
    floor (0.02 x global peak) never saw a gap there -- silently merging
    both stages into one. Built here from BASELINE_MM/cycle_arrays directly
    (not the real, gitignored file) so this specific failure mode has
    synthetic, always-available coverage independent of whether the real
    export happens to be present."""
    from conftest import BASELINE_MM, cycle_arrays

    n = 300
    s1, x1 = cycle_arrays(peak=10.0, x_perm=BASELINE_MM, amplitude=0.02, creep=0.0,
                           set_inc=0.0005, n_ramp=n, n_hold=0, n_unload=n, n_rest=0, hold=False)
    # Stop the first cycle's unload partway (a valley at ~40% of peak, well
    # above the old fixed 2%-of-global-peak floor) instead of running it to
    # conftest's own near-zero rest baseline -- this IS the failure being
    # reproduced.
    partial = n // 2
    s1, x1 = s1[: n + partial], x1[: n + partial]
    x_perm2 = BASELINE_MM + 0.0005
    s2, x2 = cycle_arrays(peak=20.0, x_perm=x_perm2, amplitude=0.02, creep=0.0,
                           set_inc=0.0005, n_ramp=n, n_hold=0, n_unload=n, n_rest=50, hold=False)

    stress = np.concatenate([s1, s2])

    cycles = segment_cycles(stress, Config())
    assert len(cycles) == 2, "the partially-relaxed valley must still separate the two stages"
    peaks = [float(stress[a : b + 1].max()) for a, b in cycles]
    assert peaks == pytest.approx([10.0, 20.0], rel=1e-3)


def test_multi_stage_is_flagged(signal):
    stress, disp = signal
    assert analyse_test(_as_test(stress, disp)).attrs["multi_stage"] is True

    flat, flat_x = multistage_signal(stages=(200.0, 200.0, 200.0))
    assert analyse_test(_as_test(flat, flat_x)).attrs["multi_stage"] is False


# ----------------------------------------------------------------------------
# Strain normalisation
# ----------------------------------------------------------------------------


def test_strain_columns_appear_only_when_h0_is_known(signal):
    """Suppressed rather than faked when the export carries no h0."""
    stress, disp = signal
    strain_cols = ["PeakStrain_pct", "PermDef_cumulative_pct", "Creep_pct"]

    with_h0 = analyse_test(_as_test(stress, disp, h0=H0_MM))
    assert all(c in with_h0.columns for c in strain_cols)
    assert with_h0["PeakStrain_pct"].iloc[-1] == pytest.approx(
        with_h0["PeakDisp_mm"].iloc[-1] / H0_MM * 100.0
    )

    without = analyse_test(_as_test(stress, disp, h0=None))
    assert not any(c in without.columns for c in strain_cols)


def test_config_h0_is_a_fallback_only(series_file, tmp_path, signal):
    """A metadata sheet wins over the configured fallback; without a sheet the
    fallback applies."""
    stress, disp = signal
    (from_meta,) = [t for t in load_tests(str(series_file), Config(h0_mm=99.0))
                    if t.label.endswith("_S1")]
    assert from_meta.h0_mm == pytest.approx(H0_MM)

    bare = write_series_workbook(
        tmp_path / "bare.xlsx", {"1": (stress, disp)}, with_metadata=False
    )
    (fallback,) = load_tests(str(bare), Config(h0_mm=1.25))
    assert fallback.h0_mm == pytest.approx(1.25)


# ----------------------------------------------------------------------------
# helpers
# ----------------------------------------------------------------------------


def _as_test(stress, disp, h0=H0_MM) -> TestData:
    return TestData(
        label="synthetic",
        displacement_mm=disp,
        stress_mpa=stress,
        source_file="synthetic",
        source_format="single",
        h0_mm=h0,
    )
